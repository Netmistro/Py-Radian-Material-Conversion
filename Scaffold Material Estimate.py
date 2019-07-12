import openpyxl as op
from openpyxl.styles import Font

# List to handle multiple files
files = ["90 x 24 x 12 Cuplock_MATERIAL LIST v2",
         "Michael York 42 x 4 x 26_MATERIAL LIST",
         "N & N Agency - 60 x 3 x 14_MATERIAL LIST"]

# Iterate through lists to handle each file
for file in range(0, (len(files))):
    fileName = files[file] + ".xlsx"
    print("Processing File..." + fileName)

    # Scaffold Materials Listing
    materialListing = {
        "Cuplok combined jack and base plate": (150.00, 125.00, 1.50),
        "Cuplok deck adaptor": (200.00, 150.00, 1.50),
        "Cuplok horizontal: 1.3m": (200.00, 175.00, 1.75),
        "Cuplok horizontal: 1.8m": (300.00, 260.00, 2.60),
        "Cuplok horizontal: 2.5m": (380.00, 330.00, 3.30),
        "Cuplok intermediate transom: 1.8m": (475.00, 0, 3.00),
        "Cuplok intermediate transom: 2.5m": (650.00, 0, 4.00),
        "Cuplok standard: 1.0m": (400.00, 325.00, 3.00),
        "Cuplok standard: 1.52m": (400.00, 325.00, 3.00),
        "Cuplok standard: 2.0m": (450.00, 360.00, 3.60),
        "Cuplok standard: 2.5m": (560.00, 450.00, 4.50),
        "Cuplok standard: 3.0m": (655.00, 550.00, 5.50),
        "Cuplok swivel face brace: 1.8 x 2.0m": (415.00, 300.00, 3.00),
        "Cuplok swivel face brace: 2.5 x 1.5m": (450.00, 350.00, 3.50),
        "Cuplok swivel face brace: 2.5 x 2.0m": (450.00, 350.00, 3.50),
        "Cuplok swivel face brace: 3.0 x 2.0m": (450.00, 400.00, 4.00),
        "Ladder hatch": (1000.00, 900.00, 15.00),
        "Ladder: steel 2.5m": (875.00, 750.00, 12.00),
        "Ladder: steel 3.0m": (1050.00, 960.00, 15.00),
        "Ladder: steel 3.5m": (1225.00, 1120.00, 18.00),
        "Ladder: steel 5.0m": (1750.00, 1600.00, 24.00),
        "Steel tube:  2'": (50.00, 43.00, 0.60),
        "Steel tube:  4'": (100.00, 86.00, 1.20),
        "Steel tube:  6'": (150.00, 129.00, 1.80),
        "Steel tube:  8'": (200.00, 172.00, 2.40),
        "Steel tube: 10'": (250.00, 215.00, 3.00),
        "Steel tube: 12'": (300.00, 258.00, 3.60),
        "Steel tube: 18'": (450.00, 387.00, 5.40),
        "Steel tube: 21'": (525.00, 451.50, 6.30),
        "Swivel coupler": (65.00, 45.00, 0.30),
        "Base plate": (55.00, 35.00, 0.30),
        "Double coupler": (60.00, 40.00, 0.30),
        "Ladder: orange painted steel 3.0m": (1050.00, 960.00, 15.00),
        "Putlog coupler": (60.00, 40.00, 0.30),
        "Sleeve coupler": (60.00, 40.00, 0.30),
        "Sole pads": (60.00, 40.00, 0.50),
        "Timber board:  1'": (30.00, 20.00, 0.25),
        "Timber board:  2'": (60.00, 40.00, 0.50),
        "Timber board:  4'": (135.00, 100.00, 1.25),
        "Timber board:  6'": (200.00, 150.00, 2.00),
        "Timber board:  8'": (250.00, 200.00, 2.50),
        "Timber board: 10'": (300.00, 250.00, 3.00),
        "Timber board: 13'": (325.00, 275.00, 3.25),
        "Timber board:  2' - 225mm x 38mm": (60.00, 40.00, 0.50),
        "Timber board:  4' - 225mm x 38mm": (135.00, 100.00, 1.25),
        "Timber board:  6' - 225mm x 38mm": (200.00, 150.00, 2.00),
        "Timber board:  8' - 225mm x 38mm": (250.00, 200.00, 2.504),
        "Timber board:  10' - 225mm x 38mm": (300.00, 250.00, 3.00),
        "Timber board:  12' - 225mm x 38mm": (325.00, 275.00, 3.25),
        "Timber board:  13' - 225mm x 38mm": (325.00, 275.00, 3.25),
        "Toe board clip": (60.00, 40.00, 0.3),
    }

    # Load Workbook
    wb = op.load_workbook(fileName)

    # Create new worksheet and check if it already exists before creating
    try:
        if "Priced Sheet" in wb.sheetnames:
            print("The pricing sheet already exists, in - " + fileName)
        else:
            print("Deleting old price sheet and creating new one!")
            wb.create_sheet("Priced Sheet")
    except Exception as e:
        print(e)

    # Create names for each sheet so they can be referred to directly
    Sheet1 = wb["Sheet"]
    Sheet2 = wb["Priced Sheet"]

    # Writing Headers to the Sheet
    headers = ["Item", "Material Description", "Quantity", "Sale Unit Price", "Sale Price", "Sale Used Price",
               "Used Price", "Rental Unit Price", "Rental Price", "Weight (lbs)"]
    try:
        i = 1
        for item in headers:
            Sheet2.cell(row=1, column=i, value=item)
            i += 1
    except Exception as e:
        print(e + "Cannot write headers!")

    # Bold Headers
    for p in range(1, 11):
        boldHeadersText = Sheet2.cell(row=1, column=p)
        boldHeadersText.font = Font(bold=True, name="Arial", size=10)


    # Search for last row of data in Sheet1
    def sheet_last_row():
        for x in range(1, 45):
            if Sheet1.cell(row=x, column=1).value == "Extra Materials":
                return x


    # Copy some values from sheet1 to sheet2
    sheet_last_row()
    for item in range(12, sheet_last_row()):
        x = item - 10
        materialDescription = Sheet1.cell(row=item, column=4).value
        materialQuantity = Sheet1.cell(row=item, column=10).value
        materialWeight = Sheet1.cell(row=item, column=13).value
        materialWeight = float(materialWeight.rstrip(" lbs"))
        Sheet2.cell(row=x, column=1, value=item - 11)
        Sheet2.cell(row=x, column=2, value=materialDescription)
        Sheet2.cell(row=x, column=3, value=materialQuantity)
        Sheet2.cell(row=x, column=10, value=materialWeight).number_format = '0.00'
        if Sheet2.cell(row=x, column=2).value in materialListing:
            saleNewPrice = Sheet2.cell(row=x, column=4, value=materialListing[materialDescription][0])
            saleUsedPrice = Sheet2.cell(row=x, column=6, value=materialListing[materialDescription][1])
            rentalUnitPrice = Sheet2.cell(row=x, column=8, value=materialListing[materialDescription][2])

    # Insert Formulae into cells - Multiplication of Quantity and Unit Price
    try:
        for j in range(2, (x + 1)):
            salePriceFormula = "=" + "C" + str(j) + "*" + "D" + str(j)
            usedPriceFormula = "=" + "C" + str(j) + "*" + "F" + str(j)
            rentalPriceFormula = "=" + "C" + str(j) + "*" + "H" + str(j)

            Sheet2.cell(row=j, column=5, value=str(salePriceFormula))
            Sheet2.cell(row=j, column=7, value=str(usedPriceFormula))
            Sheet2.cell(row=j, column=9, value=str(rentalPriceFormula))
    except Exception as e:
        print(e)

    # Determine final row to write the sums
    finalRowForSum = x + 3

    try:
        # Total quantities of each column & format each output as well
        Sheet2.cell(row=finalRowForSum, column=5,
                    value=str("=SUM(" + "E2" + ":" + "E" + str(x + 1) + ")")).number_format = '"$"#,##0_);("$"#,##0)'
        Sheet2.cell(row=finalRowForSum, column=7,
                    value=str("=SUM(" + "G2" + ":" + "G" + str(x + 1) + ")")).number_format = '"$"#,##0_);("$"#,##0)'
        Sheet2.cell(row=finalRowForSum, column=9,
                    value=str("=SUM(" + "I2" + ":" + "I" + str(x + 1) + ")")).number_format = '"$"#,##0_);("$"#,##0)'
        Sheet2.cell(row=finalRowForSum, column=10,
                    value=str("=SUM(" + "J2" + ":" + "J" + str(x + 1) + ")")).number_format = '0.00'
    except Exception as e:
        print(e)

    # Save changes to workbook
    try:
        wb.save(fileName)
        wb.close()
    except Exception as e:
        print(e)
